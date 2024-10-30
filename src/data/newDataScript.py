from openpyxl import *
import requests, config, json

# ws contains the main sheet (MasterList)
wb = load_workbook('MovieMovieMovies.xlsx')
ws = wb.active

# indexes for columns of excel sheet
titleIndex = 0
yearIndex = 9
plot = 12
poster = plot + 1
actors = poster + 1
director = actors + 1
ratings = director + 1
boxoffice = ratings + 1
rated = boxoffice + 1
runtime = rated + 1
provider = runtime + 1
budget = provider + 1
tmdbid = budget + 1
recos = tmdbid + 1
rt = recos + 1
imdb = rt + 1
metacritic = imdb + 1
trailer = metacritic + 1

apikey = config.apikey

for index, row in enumerate(ws.iter_rows(values_only=True)):
    if index >= 1:
        tmdbcode = ws[index + 1][tmdbid].value

        videos = requests.get(f'https://api.themoviedb.org/3/movie/{tmdbcode}/videos?api_key={config.tmdbkey}').json()
        trailers = [result for result in videos['results'] if result['type'] == 'Trailer']
        # If there's more than one trailer, prioritize the official one
        if len(trailers) > 0:
          selected_trailer = next((trailer for trailer in trailers if trailer['official']), trailers[0])
        elif len(videos['results']) > 0:
           selected_trailer = videos['results'][0]
        else:
           selected_trailer = ''
        ws[index + 1][trailer].value = f'https://www.youtube.com/watch?v={selected_trailer["key"]}' if selected_trailer != '' else ''

        r = json.loads(ws[index + 1][ratings].value)
        rotten = [rating['Value'] for rating in r if rating['Source'] == 'Rotten Tomatoes']
        ws[index + 1][rt].value = rotten[0] if len(rotten) > 0 else 'N/A'
        internationalMovieDB = [rating['Value'] for rating in r if rating['Source'] == 'Internet Movie Database']
        ws[index + 1][imdb].value = internationalMovieDB[0] if len(internationalMovieDB) > 0 else 'N/A'
        mc = [rating['Value'] for rating in r if rating['Source'] == 'Metacritic']
        ws[index + 1][metacritic].value = mc[0] if len(mc) > 0 else 'N/A'
        print(ws[index + 1][titleIndex].value)


wb.save('MovieMovieMovies.xlsx')
