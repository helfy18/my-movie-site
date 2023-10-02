from openpyxl import *
import requests, sys, config

# ws contains the main sheet (MasterList)
wb = load_workbook('MovieMovieMovies.xlsx')
ws = wb.active

# indexes for columns of excel sheet
title = 0
year = 9
plot = 11
poster = 12
actors = 13
director = 14
ratings = 15
boxoffice = 16
rated = 17
runtime = 18
provider = 19
budget = 20

for index, row in enumerate(ws.iter_rows(values_only=True)):
    if index >= 1:
        # skip entries already filled, comment out if full update required
        if ws[index + 1][plot].value:
            continue

        # title and year for search
        t = ws[index + 1][title].value
        y = ws[index + 1][year].value
        # Special Cases
        if t == "The Black Phone":
            y = '2021'
        if t == "Monty Python's Life of Brian":
            t = "Life of Brian"
        if t == "The Lion King 1 1/2":
            t = "The Lion King 3"
        if "Batman v Superman" in t:
            t = "Batman v Superman"
        if t == "Mom and Dad":
            y = '2017'
        if t == "Aladdin 2: The Return of Jafar":
            t = "The Return of Jafar"
        if t == "Fant4stic":
            t = "Fantastic Four"
        if "Kangaroo Jack:" in t and y == 2004:
            t = "Kangaroo Jack 2"
        if t == "Cyrano":
            y = '2021'
        if t == 'Blades of Glory':
            y = '2006'
        if "&" in t:
            t = t.replace("&", "%26")
        if t == "Tiptoes":
            y = '2002'
        if t == 'Glass Onion: A Knives Out Mystery':
            t = 'Glass Onion'
        if t == 'Marcel the Shell with Shoes On':
            y = '2021'
        # submit api request

        if t == 'Blades of Glory':
            y = '2007'
        if t == "Tiptoes":
            y = '2003'

        m2 = requests.get(f'https://api.themoviedb.org/3/search/movie?api_key={config.tmdbkey}&query={t}&year={y}').json()
        tmdbcode = m2["results"][0]["id"]
        if m2["results"][0]["original_title"] == 'X-Men: First Class 35mm Special':
            tmdbcode = m2['results'][1]['id']
        if m2['results'][0]['original_title'] == "What's Your Name?":
            tmdbcode = m2['results'][1]['id']

        url = f'https://api.themoviedb.org/3/movie/{tmdbcode}/rating'
        headers = {'Content-Type': 'application/json;charset=utf8', 'Authorization': f'{config.tmdbtoken}'}
        value = round((ws[index + 1][1].value)/5)/2
        if value == 0.0:
            value = 0.5
        data = {"value": value}

        response = requests.post(url, headers=headers, json=data).json()
        print(t, y, response, url, headers, data)
        if response['success'] == False:
            raise Exception("BAD REQUEST")