from openpyxl import *
import requests, config, json, time

# ws contains the main sheet (MasterList)
wb = load_workbook('MovieMovieMovies.xlsx')
ws = wb.active

# indexes for columns of excel sheet
titleIndex = 0
yearIndex = 9
plot = 12
poster = plot + 1
actors = poster + 1
director = actors + 1
ratings = director + 1
boxoffice = ratings + 1
rated = boxoffice + 1
runtime = rated + 1
provider = runtime + 1
budget = provider + 1
tmdbid = budget + 1
recos = tmdbid + 1
rt = recos + 1
imdb = rt + 1
metacritic = imdb + 1
trailer = metacritic + 1
millis = trailer + 1

apikey = config.apikey

currentTime = round(time.time() * 1000)

for index, row in enumerate(ws.iter_rows(values_only=True)):
    if index == 950:
       apikey = config.apikey2
    if index >= 1:
        # skip entries already filled, comment out if full update required
        # if ws[index + 1][plot].value:
        #     continue

        # title and year for search
        title = ws[index + 1][titleIndex].value
        year = ws[index + 1][yearIndex].value
        
        if not ws[index + 1][tmdbid].value:
            search = requests.get(f'https://api.themoviedb.org/3/search/movie?api_key={config.tmdbkey}&query={title}&year={year}').json()
            path = search['results'][0]['poster_path']
            tmdbcode = search["results"][0]["id"]
            ws[index + 1][poster].value = f'https://image.tmdb.org/t/p/w500{path}'
            ws[index + 1][tmdbid].value = tmdbcode

            ws[index + 1][millis].value = str(currentTime)
        else:
            tmdbcode = ws[index + 1][tmdbid].value

        if not ws[index + 1][actors].value:
            castAndCrew = requests.get(f'https://api.themoviedb.org/3/movie/{tmdbcode}/credits?api_key={config.tmdbkey}').json()
            actorString = ""
            count = 1
            for actor in castAndCrew["cast"]:
                if count < 8:
                    actorString = f'{actorString}{actor["name"]}, '
                    count = count + 1
                else:
                    actorString = f'{actorString}{actor["name"]}'
                    break
            if actorString and actorString[-2] == ",":
                actorString = actorString[:-2]
            print(title, actorString)
            if actorString:
                ws[index + 1][actors].value = actorString
            else:
                ws[index + 1][actors].value = "N/A"
            
            ws[index + 1][director].value = ', '.join(credit["name"] for credit in castAndCrew["crew"] if credit["job"] == "Director")
        
        movieInfo = requests.get(f'https://api.themoviedb.org/3/movie/{tmdbcode}?api_key={config.tmdbkey}').json()
        try:
            boxofficeTotal = movieInfo['revenue']
        except KeyError:
            boxofficeTotal = 'N/A'
        imdbid = movieInfo["imdb_id"]
        ws[index + 1][plot].value = movieInfo["overview"]
        ws[index + 1][boxoffice].value = f"{boxofficeTotal:,}"
        ws[index + 1][budget].value = f"{movieInfo['budget']:,}"
        ws[index + 1][runtime].value = f"{movieInfo['runtime']:,}"
        
        providers = requests.get(f'https://api.themoviedb.org/3/movie/{tmdbcode}/watch/providers?api_key={config.tmdbkey}').json()
        if providers['results'] and 'CA' in providers['results']:
            ws[index + 1][provider].value = json.dumps(providers['results']['CA'])
        else:
            ws[index + 1][provider].value = "{" + "}"

        recommendations = requests.get(f'https://api.themoviedb.org/3/movie/{tmdbcode}/recommendations?api_key={config.tmdbkey}').json()
        ws[index + 1][recos].value = str([item['id'] for item in recommendations['results']])

        videos = requests.get(f'https://api.themoviedb.org/3/movie/{tmdbcode}/videos?api_key={config.tmdbkey}').json()
        trailers = [result for result in videos['results'] if result['type'] == 'Trailer']
        # If there's more than one trailer, prioritize the official one
        if len(trailers) > 0:
          selected_trailer = next((trailer for trailer in trailers if trailer['official']), trailers[0])
        elif len(videos['results']) > 0:
           selected_trailer = videos['results'][0]
        else:
           selected_trailer = ''
        ws[index + 1][trailer].value = f'https://www.youtube.com/embed/{selected_trailer["key"]}' if selected_trailer != '' else ''

        url = f'https://api.themoviedb.org/3/movie/{tmdbcode}/rating'
        headers = {'Content-Type': 'application/json;charset=utf8', 'Authorization': f'Bearer {config.tmdbtoken}'}
        if ws[index + 1][1].value:
            value = round((ws[index + 1][1].value)/5)/2
            if value == 0.0:
                value = 0.5
            data = {"value": value}
            response = requests.post(url, headers=headers, json=data).json()

        # OMDB SECTION
        omdb = requests.get(f'http://www.omdbapi.com/?apikey={apikey}&i={imdbid}&type=movie').json()
        rotten = [rating['Value'] for rating in omdb["Ratings"] if rating['Source'] == 'Rotten Tomatoes']
        ws[index + 1][rt].value = rotten[0] if len(rotten) > 0 else 'N/A'
        internationalMovieDB = [rating['Value'] for rating in omdb["Ratings"] if rating['Source'] == 'Internet Movie Database']
        ws[index + 1][imdb].value = internationalMovieDB[0] if len(internationalMovieDB) > 0 else 'N/A'
        mc = [rating['Value'] for rating in omdb["Ratings"] if rating['Source'] == 'Metacritic']
        ws[index + 1][metacritic].value = mc[0] if len(mc) > 0 else 'N/A'

        ws[index + 1][ratings].value = json.dumps(omdb["Ratings"])
        ws[index + 1][rated].value = omdb["Rated"]

        print(title, year, index)
        
wb.save('MovieMovieMovies.xlsx')
